import os
import shutil
import textract
import xlrd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
import re
import threading
from pdfminer.pdfparser import PDFParser, PSSyntaxError
from pdfminer.pdfdocument import PDFDocument
from docx import Document
import pytesseract
from pdf2image import convert_from_path
import PyPDF2
import platform
from PIL import Image
from docx.opc.exceptions import PackageNotFoundError
if platform.system() == "Windows":
    from win32com.client import constants
    import win32com.client as win32

# __________________TODO__________________
#     

# Список обрабатываемых textract'ом типов документов
file_types = ('.docx', '.xlsx', '.ppt', '.odt')
image_files = ('.jpg', '.png', '.jpeg', '.bmp', '.tif')


def timeout_handler():
    raise Exception("Timeout Exception")

def move_subfolder_contents(folder_path):
    for root, dirs, files in os.walk(folder_path):
        if root != folder_path:  # Исключаем основную папку
            for file in files:
                src = os.path.join(root, file)
                dst = os.path.join(folder_path, file)
                if not os.path.exists(dst):
                    shutil.move(src, dst)


def convert_xls_to_xlsx(input_files):
    for filename in os.listdir(input_files):
        if filename.endswith('.xls'):
            xls_path = os.path.join(input_files, filename)
            wb = xlrd.open_workbook(xls_path)

            xlsx_path = os.path.join(input_files, filename[:-4] + '.xlsx')
            new_wb = Workbook()

            ws = new_wb.active

            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)

                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        col_letter = get_column_letter(col + 1)
                        cell_value = sheet.cell_value(row, col)
                        ws[f"{col_letter}{row + 1}"] = cell_value

            new_wb.save(xlsx_path)
            os.remove(xls_path)


def convert_doc_to_docx(input_files):
    for filename in os.listdir(input_files):
        if filename.endswith('.doc'):
            input_path = os.path.join('D:\\Projects\\tsiars_gpt\\input_files', filename)
            word = win32.gencache.EnsureDispatch('Word.Application')
            doc = word.Documents.Open(input_path)
            print(doc)
            doc.Activate()

            new_file_abs = os.path.abspath(input_path)
            new_file_abs = re.sub(r'\.\w+$', '.docx', new_file_abs)
            word.ActiveDocument.SaveAs(
                new_file_abs, FileFormat=constants.wdFormatXMLDocument
            )
            doc.Close(False)
            os.remove(input_path)


def metadata_extracter(input_files, output_txt):
    df = pd.DataFrame(columns=['Meta', 'Path'])
    for filename in os.listdir(input_files):
        if filename.endswith(".docx"):
            try:
                doc = Document(os.path.join(input_files, filename))
                metadata = doc.core_properties
            except PackageNotFoundError:
                print("Package not found error occurred!")
            metadata_dict = {"Title": metadata.title,
                             "Author": metadata.author,
                             "Subject": metadata.subject,
                             "Keywords": metadata.keywords,
                             "Category": metadata.category,
                             "Comments": metadata.comments}
            print(metadata_dict)
            df = df._append({'Meta': metadata_dict,
                             'Path': os.path.join(output_txt, f'{filename[:-5]}.txt')},
                            ignore_index=True)

        elif filename.endswith(".pdf"):
            with open(os.path.join(input_files, filename), 'rb') as pdf_file:
                parser = PDFParser(pdf_file)
                try:
                    doc = PDFDocument(parser)
                except PSSyntaxError:
                    doc = None
                try:
                    metadata_dict = doc.info[0]
                except:
                    metadata_dict = {}
                print(metadata_dict)

                df = df._append({'Meta': metadata_dict,
                                 'Path': os.path.join(output_txt, f'{filename[:-5]}.txt')},
                                ignore_index=True)

        elif filename.endswith(".xlsx") or filename.endswith(".xlsm") or filename.endswith(
                ".xltx") or filename.endswith(".xltm"):
            wb = openpyxl.load_workbook(os.path.join(input_files, filename))
            metadata_dict = {}
            try:
                wb = openpyxl.load_workbook(os.path.join(input_files, filename))
                metadata_dict['authors'] = wb.properties.creator
                metadata_dict['theme'] = wb.properties.title
                metadata_dict['tags'] = wb.properties.keywords
                metadata_dict['category'] = wb.properties.category
                metadata_dict['comments'] = wb.properties.description
            except TypeError:
                print("TypeError occurred while loading the Excel file. Continuing with the code.")
                wb = None
            print(metadata_dict)
            df = df._append({'Meta': metadata_dict,
                             'Path': os.path.join(output_txt, f'{filename[:-5]}.txt')},
                            ignore_index=True)

    df.to_csv('dataframe.csv', index=False)


def textract_converter(input_files, output_txt):
    for filename in os.listdir(input_files):
        print(f'Now Converting {filename}')
        file_path = os.path.join(input_files, filename)
        a = False
        if filename.endswith('.pdf'):
            with open(f'input_files/{filename}', 'rb') as file:
                a = False
                try:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text = page.extract_text()
                        if text:
                            a = True
                except Exception:     
                    a = None
            if a == False:
                pages = convert_from_path(os.path.join(input_files, filename), 400)
                text = ""
                try:
                    timer = threading.Timer(600, timeout_handler)
                    # Start the timer
                    timer.start()
                    for pageNum, imgBlob in enumerate(pages):
                        text += pytesseract.image_to_string(imgBlob, lang='rus') + '\n'
                    timer.cancel()
                except Exception as e:
                    print("Error: ", e)
                finally:               
                    new_filename = os.path.splitext(filename)[0] + '.txt'
                    with open(os.path.join(output_txt, new_filename), 'w', encoding='utf-8') as f:
                        f.write(text)
            else:
                try:
                    text = textract.process(os.path.join(input_files, filename)).decode('utf-8')
                except:
                    print('Error on textract.process')
                    text = 'Error Text'
                new_filename = os.path.splitext(filename)[0] + '.txt'
                with open(os.path.join(output_txt, new_filename), 'w', encoding='utf-8') as f:
                    f.write(text)

        if filename.endswith(file_types):
            try:
                text = textract.process(os.path.join(input_files, filename)).decode('utf-8')
            except:
                print('Error on textract.process')
                text = 'Error Text'
            new_filename = os.path.splitext(filename)[0] + '.txt'
            with open(os.path.join(output_txt, new_filename), 'w', encoding='utf-8') as f:
                f.write(text)

        if filename.endswith('.html'):
            file_path = os.path.join(input_files, filename)
            output_path = os.path.join(output_txt, filename) 
            with open(file_path, 'rb') as html_file:
                html_content = html_file.read()
                soup = BeautifulSoup(html_content, 'html.parser')
                text_content = soup.get_text()
                
                new_file_path = os.path.splitext(output_path)[0] + '.txt'
            with open(new_file_path, 'w', encoding='utf-8') as text_file:
                text_file.write(text_content)       

        if filename.endswith(image_files):
            image = Image.open(os.path.join(input_files, filename))
            text = pytesseract.image_to_string(image, lang='rus')
            new_filename = os.path.splitext(filename)[0] + '.txt'
            with open(os.path.join(output_txt, new_filename), 'w', encoding='utf-8') as f:
                f.write(text)
                
        os.remove(file_path)
            

        if filename.endswith('.html'):
            file_path = os.path.join(input_files, filename)
            with open(file_path, 'rb') as html_file:
                html_content = html_file.read()

                soup = BeautifulSoup(html_content, 'html.parser')
                text = soup.get_text()

                new_filename = os.path.splitext(filename)[0] + '.txt'
                with open(os.path.join(output_txt, new_filename), 'w', encoding='utf-8') as f:
                    f.write(text)


def lines_editor(output_txt):
    for filename in os.listdir(output_txt):
        if filename.endswith('.txt'):
            filepath = os.path.join(output_txt, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            lines = [line for line in lines if line.strip()]
            with open(filepath, 'w', encoding='utf-8') as f:
                f.writelines(lines)

            with open(filepath, 'r', encoding='utf-8') as file:
                text = file.read()
            text = re.sub(r'([а-я])([А-Я])', r'\1 \2', text)
            text = re.sub(r'(\w+)-\n(\w+)', r'\1\2', text)
            text = re.sub(r'(?<=\w)– ', '', text)
            text = re.sub(r'(?<=\w)- ', '', text)
            text = text.replace(" ", " ")
            text = text.replace("�", "")
            text = text.replace("", " ")
            text = re.sub(r'(?<=\b\w)\s(?=\w\b)', '', text)
            # text = re.sub(r'[^\w\d,.\- ]', '', text)

            with open(filepath, 'w', encoding='utf-8') as file:
                file.write(text)            